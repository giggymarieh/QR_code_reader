import openpyxl
import pandas as pd

class ExcelWriter:
    workbook = openpyxl.Workbook()
    input_file_path : str
    output_file_path : str
    def __init__(self, input_file_path, output_file_path):
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path

    def append_row_to_excel(self, value, timestamp, sheet_name):
        try:
            # Load the existing Excel workbook if it exists
            workbook = openpyxl.load_workbook(self.output_file_path)
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            workbook = openpyxl.Workbook()

        # Check if the specified sheet already exists
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            # Create a new sheet if it doesn't exist
            sheet = workbook.create_sheet(title=sheet_name)

        # Calculate the index for the new row (after the last row)
        insert_index = sheet.max_row + 1

        # Create a DataFrame with the new row data
        data = {'Value': value, 'Timestamp': timestamp}
        df_row = pd.DataFrame([data])

        # Convert the DataFrame row to a list
        row_values = df_row.values.tolist()[0]

        # Insert a new row at the calculated index
        sheet.insert_rows(insert_index)

        # Populate the cells in the new row with data
        for col, value in enumerate(row_values, start=1):
            sheet.cell(row=insert_index, column=col, value=value)

        # Save the modified workbook
        workbook.save(self.output_file_path)

        # Close the workbook
        workbook.close()

    def get_sheet_names(self) -> list[str]:
        if self.input_file_path:
            wb = openpyxl.load_workbook(self.input_file_path)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        
    def get_sessions(self, selected_sheet: str) -> list[str]:
        if selected_sheet:
            wb = openpyxl.load_workbook(self.input_file_path)
            sheet = wb[selected_sheet]
            rows = sheet.iter_rows(values_only=True)
            row_values = [", ".join(map(str, row)) for row in rows]
            wb.close() 
            return row_values